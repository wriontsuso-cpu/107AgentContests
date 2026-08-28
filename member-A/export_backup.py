#!/usr/bin/env python3
"""将现有公开采集结果整理为文档 / 表格 / PDF 备份（非登录可获取信息）。"""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from fpdf import FPDF

ROOT = Path(__file__).resolve().parent
INPUT_JSON = ROOT / "output" / "student_resources.json"
BACKUP_ROOT = ROOT / "backup" / "非登录可获取信息"


def load_articles() -> tuple[dict, list[dict]]:
    payload = json.loads(INPUT_JSON.read_text(encoding="utf-8"))
    articles = payload.get("articles") or []
    return payload, articles


def access_label(item: dict) -> str:
    tags = item.get("tags") or []
    text = " ".join(
        [
            item.get("title", ""),
            item.get("summary", ""),
            item.get("how_to", ""),
            " ".join(tags),
        ]
    )
    if "需登录" in tags or "需登录" in text or "统一身份认证" in text:
        return "公开入口说明（目标系统需登录）"
    return "公开可访问"


def safe_filename(name: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", name).strip()
    return cleaned[:60] or "未命名"


def ensure_dirs() -> dict[str, Path]:
    paths = {
        "root": BACKUP_ROOT,
        "docs": BACKUP_ROOT / "文档",
        "tables": BACKUP_ROOT / "表格",
        "pdf": BACKUP_ROOT / "PDF",
        "by_cat": BACKUP_ROOT / "文档" / "按栏目",
    }
    for path in paths.values():
        path.mkdir(parents=True, exist_ok=True)
    return paths


def write_readme(paths: dict[str, Path], payload: dict, articles: list[dict]) -> None:
    generated = payload.get("generated_at", "")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    curated = sum(1 for a in articles if a.get("kind") == "curated")
    crawled = len(articles) - curated
    public_n = sum(1 for a in articles if access_label(a) == "公开可访问")
    portal_n = len(articles) - public_n

    text = f"""# 非登录可获取信息 · 备份包

- 备份生成时间：{now}
- 原始采集时间：{generated}
- 条目总数：{len(articles)}（精选 {curated} + 公开爬取 {crawled}）
- 公开可访问条目：{public_n}
- 公开入口说明（系统本身需登录）：{portal_n}

## 说明

本备份整理自当前爬虫与精选目录的**已采集结果**，标记为「非登录可获取信息」：

1. 网页标题、链接、摘要、栏目等来自**无需登录即可打开**的公开页面，或精选目录中的公开入口说明。
2. 其中少量条目是「需登录系统」的**入口与用法说明**（如青春科大、境外交流系统），条目本身公开，但系统内部活动列表仍需你自行登录查看。
3. 不包含任何账号密码，也不包含登录后才能看到的私有数据。

## 目录结构

- `文档/`：总览 Markdown、按栏目分册
- `表格/`：CSV + Excel（全量与分表）
- `PDF/`：汇总 PDF（便于打印/归档）

## 建议阅读顺序

1. `文档/00_总览与使用说明.md`
2. `表格/非登录可获取信息_全量.xlsx`
3. `PDF/非登录可获取信息_汇总.pdf`
"""
    (paths["root"] / "README.md").write_text(text, encoding="utf-8")


def write_overview_md(paths: dict[str, Path], payload: dict, articles: list[dict]) -> Path:
    cats = Counter(a.get("category") or "未分类" for a in articles)
    sources = Counter(a.get("source") or "未知" for a in articles)
    lines = [
        "# 非登录可获取信息 · 总览",
        "",
        f"- 原始采集时间：{payload.get('generated_at', '')}",
        f"- 条目总数：{len(articles)}",
        "",
        "## 按栏目",
        "",
    ]
    for cat, n in cats.most_common():
        lines.append(f"- {cat}: {n}")

    lines.extend(["", "## 按来源（前 20）", ""])
    for src, n in sources.most_common(20):
        lines.append(f"- {src}: {n}")

    lines.extend(["", "## 精选资源（含申请说明）", ""])
    for a in articles:
        if a.get("kind") != "curated":
            continue
        lines.append(f"### {a.get('title', '')}")
        lines.append(f"- 栏目：{a.get('category', '')}")
        lines.append(f"- 访问：{access_label(a)}")
        lines.append(f"- 链接：{a.get('url', '')}")
        if a.get("cost"):
            lines.append(f"- 费用：{a.get('cost')}")
        if a.get("how_to"):
            lines.append(f"- 获取方式：{a.get('how_to')}")
        if a.get("summary"):
            lines.append(f"- 简介：{a.get('summary')}")
        lines.append("")

    target = paths["docs"] / "00_总览与使用说明.md"
    target.write_text("\n".join(lines), encoding="utf-8")
    return target


def write_category_mds(paths: dict[str, Path], articles: list[dict]) -> None:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for a in articles:
        grouped[a.get("category") or "未分类"].append(a)

    index_lines = ["# 按栏目文档索引", ""]
    for cat in sorted(grouped.keys()):
        items = grouped[cat]
        fname = f"{safe_filename(cat)}.md"
        path = paths["by_cat"] / fname
        lines = [f"# {cat}", "", f"共 {len(items)} 条", ""]
        for i, a in enumerate(items, 1):
            lines.append(f"## {i}. {a.get('title', '')}")
            lines.append(f"- 访问属性：{access_label(a)}")
            lines.append(f"- 来源：{a.get('source', '')}")
            lines.append(f"- 链接：{a.get('url', '')}")
            if a.get("published_at"):
                lines.append(f"- 日期：{a.get('published_at')}")
            if a.get("cost"):
                lines.append(f"- 费用：{a.get('cost')}")
            if a.get("how_to"):
                lines.append(f"- 获取方式：{a.get('how_to')}")
            tags = a.get("tags") or []
            if tags:
                lines.append(f"- 标签：{' | '.join(tags)}")
            summary = (a.get("summary") or "").strip()
            if summary:
                lines.append(f"- 摘要：{summary[:500]}")
            lines.append("")
        path.write_text("\n".join(lines), encoding="utf-8")
        index_lines.append(f"- [{cat}](按栏目/{fname})（{len(items)}）")

    (paths["docs"] / "01_按栏目索引.md").write_text("\n".join(index_lines), encoding="utf-8")


def write_full_markdown(paths: dict[str, Path], articles: list[dict]) -> Path:
    lines = ["# 非登录可获取信息 · 全量清单", ""]
    for i, a in enumerate(articles, 1):
        lines.append(f"## {i}. [{a.get('category', '')}] {a.get('title', '')}")
        lines.append(f"- 访问属性：{access_label(a)}")
        lines.append(f"- 来源：{a.get('source', '')}")
        lines.append(f"- URL：{a.get('url', '')}")
        if a.get("published_at"):
            lines.append(f"- 发布：{a.get('published_at')}")
        if a.get("cost"):
            lines.append(f"- 费用：{a.get('cost')}")
        if a.get("how_to"):
            lines.append(f"- 获取：{a.get('how_to')}")
        summary = (a.get("summary") or "").strip()
        if summary:
            lines.append(f"- 摘要：{summary[:800]}")
        lines.append("")
    target = paths["docs"] / "02_全量清单.md"
    target.write_text("\n".join(lines), encoding="utf-8")
    return target


def row_dict(a: dict) -> dict:
    return {
        "访问属性": access_label(a),
        "类型": "精选目录" if a.get("kind") == "curated" else "公开爬取",
        "栏目": a.get("category", ""),
        "标题": a.get("title", ""),
        "链接": a.get("url", ""),
        "来源": a.get("source", ""),
        "发布日期": a.get("published_at", ""),
        "费用": a.get("cost", ""),
        "获取方式": a.get("how_to", ""),
        "标签": " | ".join(a.get("tags") or []),
        "相关分": a.get("relevance_score", ""),
        "摘要": (a.get("summary") or "")[:1000],
        "采集时间": a.get("crawled_at", ""),
    }


def write_csv(paths: dict[str, Path], articles: list[dict]) -> Path:
    import csv

    rows = [row_dict(a) for a in articles]
    fields = list(rows[0].keys()) if rows else []
    target = paths["tables"] / "非登录可获取信息_全量.csv"
    with target.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    return target


def write_excel(paths: dict[str, Path], articles: list[dict]) -> Path:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    # 总览 sheet
    overview = wb.active
    overview.title = "总览"
    cats = Counter(a.get("category") or "未分类" for a in articles)
    overview.append(["非登录可获取信息备份"])
    overview.append(["生成时间", datetime.now().isoformat(timespec="seconds")])
    overview.append(["条目总数", len(articles)])
    overview.append([])
    overview.append(["栏目", "数量"])
    for cat, n in cats.most_common():
        overview.append([cat, n])

    # 全量 sheet
    full = wb.create_sheet("全量条目")
    rows = [row_dict(a) for a in articles]
    headers = list(rows[0].keys()) if rows else []
    full.append(headers)
    for cell in full[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        full.append([row[h] for h in headers])
    for col in range(1, len(headers) + 1):
        full.column_dimensions[get_column_letter(col)].width = 18
    full.column_dimensions["D"].width = 40
    full.column_dimensions["E"].width = 40
    full.column_dimensions["L"].width = 50
    for row in full.iter_rows(min_row=2, max_row=min(full.max_row, 5000)):
        for cell in row:
            cell.alignment = wrap

    # 精选 sheet
    curated = wb.create_sheet("精选目录")
    curated_rows = [row_dict(a) for a in articles if a.get("kind") == "curated"]
    curated.append(headers)
    for cell in curated[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in curated_rows:
        curated.append([row[h] for h in headers])

    # 分栏目 sheets（避免 sheet 名过长）
    grouped: dict[str, list[dict]] = defaultdict(list)
    for a in articles:
        grouped[a.get("category") or "未分类"].append(a)
    used_names: set[str] = {"总览", "全量条目", "精选目录"}
    for cat, items in sorted(grouped.items(), key=lambda x: (-len(x[1]), x[0])):
        name = safe_filename(cat)[:28]
        base = name
        i = 1
        while name in used_names:
            name = f"{base[:24]}_{i}"
            i += 1
        used_names.add(name)
        ws = wb.create_sheet(name)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for a in items:
            row = row_dict(a)
            ws.append([row[h] for h in headers])

    target = paths["tables"] / "非登录可获取信息_全量.xlsx"
    wb.save(target)
    return target


class ChinesePDF(FPDF):
    def footer(self) -> None:
        self.set_y(-15)
        self.set_x(self.l_margin)
        self.set_font("cn", size=9)
        self.cell(0, 10, f"Page {self.page_no()} | USTC public info backup", align="C")


def find_chinese_font() -> Path:
    candidates = [
        Path(r"C:\Windows\Fonts\simhei.ttf"),
        Path(r"C:\Windows\Fonts\simkai.ttf"),
        Path(r"C:\Windows\Fonts\simfang.ttf"),
        Path(r"C:\Windows\Fonts\msyh.ttf"),
        Path(r"C:\Windows\Fonts\msyh.ttc"),
        Path(r"C:\Windows\Fonts\simsun.ttc"),
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError("未找到可用的中文字体（黑体/楷体/微软雅黑等）")


def write_pdf(paths: dict[str, Path], payload: dict, articles: list[dict]) -> Path:
    font_path = find_chinese_font()
    pdf = ChinesePDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_font("cn", fname=str(font_path))

    def text(size: int, content: str, h: float | None = None) -> None:
        pdf.set_x(pdf.l_margin)
        pdf.set_font("cn", size=size)
        line_h = h if h is not None else max(5, size * 0.55)
        # 去掉易导致宽度计算异常的特殊符号
        cleaned = (
            content.replace("\u00a0", " ")
            .replace("\ufeff", "")
            .replace("\u2011", "-")
            .replace("·", "-")
            .replace("—", "-")
            .replace("–", "-")
            .replace("\u200b", "")
        )
        pdf.multi_cell(w=0, h=line_h, text=cleaned, new_x="LMARGIN", new_y="NEXT")

    pdf.add_page()
    text(18, "中国科学技术大学")
    text(16, "非登录可获取信息 - 汇总备份")
    text(11, f"备份生成：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    text(11, f"原始采集：{payload.get('generated_at', '')}")
    text(11, f"条目总数：{len(articles)}")
    text(
        10,
        "说明：本 PDF 为公开可访问信息与公开入口说明的归档摘要；"
        "完整字段请以同目录表格 Excel/CSV 及 Markdown 文档为准。"
        "不含账号密码，也不含登录后私有数据。",
    )

    cats = Counter(a.get("category") or "未分类" for a in articles)
    text(14, "一、栏目统计")
    for cat, n in cats.most_common():
        text(11, f"- {cat}: {n}")

    curated = [a for a in articles if a.get("kind") == "curated"]
    pdf.add_page()
    text(14, "二、精选资源目录")
    for a in curated:
        text(11, a.get("title", "") or "(无标题)")
        text(9, f"栏目：{a.get('category', '')} | {access_label(a)}")
        text(9, f"链接：{a.get('url', '')}")
        if a.get("cost"):
            text(9, f"费用：{a.get('cost')}")
        if a.get("how_to"):
            text(9, f"获取：{a.get('how_to')}")
        if a.get("summary"):
            text(9, f"简介：{(a.get('summary') or '')[:220]}")

    grouped: dict[str, list[dict]] = defaultdict(list)
    for a in articles:
        if a.get("kind") == "curated":
            continue
        grouped[a.get("category") or "未分类"].append(a)

    pdf.add_page()
    text(14, "三、公开爬取条目摘录（每栏目最多 12 条）")
    for cat in sorted(grouped.keys()):
        items = grouped[cat][:12]
        text(12, f"[{cat}] 共 {len(grouped[cat])} 条，摘录 {len(items)} 条")
        for a in items:
            title = (a.get("title") or "")[:80]
            url = a.get("url") or ""
            text(9, f"- {title}")
            text(8, f"  {url}")

    text(11, "四、完整数据位置")
    text(10, "表格/非登录可获取信息_全量.xlsx")
    text(10, "表格/非登录可获取信息_全量.csv")
    text(10, "文档/02_全量清单.md")
    text(10, "文档/按栏目/")

    target = paths["pdf"] / "非登录可获取信息_汇总.pdf"
    pdf.output(str(target))
    return target


def write_manifest(paths: dict[str, Path], files: list[Path], articles: list[dict]) -> None:
    lines = [
        "非登录可获取信息 · 文件清单",
        "=" * 40,
        f"生成时间: {datetime.now().isoformat(timespec='seconds')}",
        f"条目数: {len(articles)}",
        "",
        "文件:",
    ]
    for path in files:
        rel = path.relative_to(paths["root"])
        size = path.stat().st_size if path.exists() else 0
        lines.append(f"- {rel} ({size} bytes)")
    (paths["root"] / "文件清单.txt").write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    if not INPUT_JSON.exists():
        raise SystemExit(f"找不到采集结果: {INPUT_JSON}，请先运行 py main.py --full --no-body")

    payload, articles = load_articles()
    paths = ensure_dirs()

    # 同步一份原始 JSON 到备份包
    raw_copy = paths["root"] / "原始数据_student_resources.json"
    raw_copy.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    write_readme(paths, payload, articles)
    overview = write_overview_md(paths, payload, articles)
    write_category_mds(paths, articles)
    full_md = write_full_markdown(paths, articles)
    csv_path = write_csv(paths, articles)
    xlsx_path = write_excel(paths, articles)
    pdf_path = write_pdf(paths, payload, articles)

    files = [
        raw_copy,
        paths["root"] / "README.md",
        overview,
        paths["docs"] / "01_按栏目索引.md",
        full_md,
        csv_path,
        xlsx_path,
        pdf_path,
    ]
    write_manifest(paths, files, articles)

    print("备份完成（非登录可获取信息）")
    print(f"目录: {paths['root']}")
    print(f"条目: {len(articles)}")
    print(f"Excel: {xlsx_path}")
    print(f"PDF:   {pdf_path}")
    print(f"文档:  {paths['docs']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
